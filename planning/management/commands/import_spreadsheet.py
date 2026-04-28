"""
planning/management/commands/import_spreadsheet.py
— Custom Django management command to import the Excel workforce planning spreadsheet.

CHANGELOG:
  v2 — Updated for the March 2026 spreadsheet revision:
    - PERSON_SHEETS updated: added 'AF' (Ali Feizollah) and 'PS' (Paul Slavin);
      removed 'HI1' and 'SRSE1' which no longer exist in the spreadsheet.
    - _import_staff updated: staff sheet now has a new 'Employee number' column
      at index 2, shifting all subsequent columns right by one. Also reads the
      new 'PCM projected till' column.
    - Everything else (projects, costs, allocations, swaps) is unchanged.

HOW MANAGEMENT COMMANDS WORK:
  Django lets you create custom CLI commands that run within the Django environment.
  They live in <app>/management/commands/<command_name>.py.
  Run with: python manage.py import_spreadsheet

IDEMPOTENCY:
  We use get_or_create() throughout, which means running the import twice won't
  create duplicate records. You can re-run it after updating the spreadsheet.
"""

from django.core.management.base import BaseCommand, CommandError
import openpyxl
import datetime

from planning.models import (
    StaffMember, StaffCost, Project, ProjectBudget,
    WorkPackage, StaffAllocation, BudgetSwap
)

# -------------------------------------------------------------------------
# PERSON_SHEETS — updated for March 2026 spreadsheet
#
# Each entry here corresponds to a tab in the spreadsheet named after a
# person's initials. The import reads each tab and creates StaffAllocation
# records for that person.
#
# Changes from v1:
#   ADDED:   'AF' (Ali Feizollah), 'PS' (Paul Slavin)
#   REMOVED: 'HI1' (sheet deleted), 'SRSE1' (sheet deleted)
# -------------------------------------------------------------------------
PERSON_SHEETS = [
    'CAG', 'SSR', 'CD', 'CJ',
    'AF',                        # NEW: Ali Feizollah
    'AG', 'AH', 'AL', 'AN', 'AZ',
    'DL', 'EC', 'EM', 'ES', 'FB', 'GL', 'LS', 'OS', 'OW',
    'MA', 'MJ', 'MS', 'NJ',
    'PS',                        # NEW: Paul Slavin
    'PR', 'SO', 'SS', 'WD',
    # 'HI1' — REMOVED: sheet no longer exists in spreadsheet
    'HI2',
    # 'SRSE1' — REMOVED: sheet no longer exists in spreadsheet
    'SRSE2', 'SRSE3',
]


def safe_float(val):
    """
    Convert a value to float, returning None if it can't be converted.
    Spreadsheet cells sometimes contain formula strings instead of computed values.
    """
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        if val.startswith('=') or val == '#N/A':
            return None
        try:
            return float(val)
        except ValueError:
            return None
    return None


def safe_date(val):
    """Convert a value to a date, handling datetime objects and None."""
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    return None


def safe_str(val):
    """Convert to string, returning empty string for None."""
    if val is None:
        return ''
    s = str(val).strip()
    # Treat 'N/A' as blank for date/optional fields
    return s


def safe_str_not_na(val):
    """Like safe_str but returns '' for 'N/A' values."""
    s = safe_str(val)
    return '' if s.upper() == 'N/A' else s


class Command(BaseCommand):
    """
    Import the eScience Lab workforce planning spreadsheet into the database.

    Usage:
        python manage.py import_spreadsheet
        python manage.py import_spreadsheet --file=/path/to/spreadsheet.xlsx
    """

    help = 'Import workforce planning data from the eScience Lab Excel spreadsheet'

    def add_arguments(self, parser):
        """Define command-line arguments."""
        parser.add_argument(
            '--file',
            type=str,
            default='esl-workforce-plan.xlsx',
            help='Path to the Excel spreadsheet file'
        )

    def handle(self, *args, **options):
        """Entry point — called when the command runs."""
        filepath = options['file']
        self.stdout.write(f"Starting import from: {filepath}")

        try:
            result = self.import_data(filepath)
            self.stdout.write(self.style.SUCCESS(f"Import complete! {result}"))
        except FileNotFoundError:
            raise CommandError(f"Spreadsheet file not found: {filepath}")
        except Exception as e:
            raise CommandError(f"Import failed: {e}")

    def import_data(self, filepath):
        """
        Core import logic — can be called from the view as well as the CLI.
        Returns a summary string.
        """
        # data_only=True reads computed cell values rather than formula strings
        wb = openpyxl.load_workbook(filepath, data_only=True)

        counts = {
            'staff': 0, 'costs': 0, 'projects': 0,
            'work_packages': 0, 'allocations': 0, 'swaps': 0
        }

        # Import in dependency order — staff and projects must exist before allocations
        counts['staff'] = self._import_staff(wb)
        counts['costs'] = self._import_costs(wb)
        counts['projects'] = self._import_projects(wb)
        self._import_project_budgets(wb)
        counts['work_packages'] = self._import_work_packages(wb)
        counts['allocations'] = self._import_allocations(wb)
        counts['swaps'] = self._import_swaps(wb)

        return (
            f"Staff: {counts['staff']}, "
            f"Costs: {counts['costs']}, "
            f"Projects: {counts['projects']}, "
            f"WPs: {counts['work_packages']}, "
            f"Allocations: {counts['allocations']}, "
            f"Swaps: {counts['swaps']}"
        )

    def _import_staff(self, wb):
        """
        Import staff members from the 'staff' sheet.

        COLUMN MAP (v2 — 'Employee number' added at index 2):
          0: Initials
          1: Name
          2: Employee number       ← NEW in v2
          3: Type
          4: Dept/Group
          5: PCM until
          6: PCM projected till    ← NEW in v2
          7: timesheets approved (incl)
          8: Notes

        In v1 the sheet had no Employee number column, so indices were:
          0: Initials, 1: Name, 2: Type, 3: Dept/Group,
          4: PCM until, 5: (blank), 6: timesheets, 7: Notes
        """
        if 'staff' not in wb.sheetnames:
            return 0

        ws = wb['staff']
        count = 0

        # iter_rows(min_row=2) skips the header row
        for row in ws.iter_rows(min_row=2, values_only=True):
            initials = safe_str(row[0])
            if not initials:
                continue

            name          = safe_str(row[1])
            employee_num  = safe_str_not_na(row[2])   # index 2 — new column
            staff_type    = safe_str(row[3])           # was index 2 in v1
            department    = safe_str(row[4])           # was index 3 in v1
            pcm_until_raw = row[5]                     # was index 4 in v1
            pcm_proj_raw  = row[6]                     # NEW: PCM projected till
            ts_raw        = row[7]                     # was index 6 in v1
            notes         = safe_str(row[8])           # was index 7 in v1

            # PCM until: could be 'N/A' (for academics) or a date
            pcm_until = None
            if pcm_until_raw and safe_str(pcm_until_raw).upper() != 'N/A':
                pcm_until = safe_date(pcm_until_raw)

            # PCM projected until: same treatment
            pcm_projected = None
            if pcm_proj_raw and safe_str(pcm_proj_raw).upper() not in ('N/A', ''):
                pcm_projected = safe_date(pcm_proj_raw)

            # Timesheets approved until: could be 'N/A' or a date
            ts_until = None
            if ts_raw and safe_str(ts_raw).upper() != 'N/A':
                ts_until = safe_date(ts_raw)

            # get_or_create: look up by 'initials'; if not found, create with defaults.
            # If found, we update the fields in case the spreadsheet has changed.
            obj, created = StaffMember.objects.get_or_create(
                initials=initials,
                defaults={
                    'name': name,
                    'employee_number': employee_num,
                    'staff_type': staff_type,
                    'department': department,
                    'pcm_until': pcm_until,
                    'pcm_projected_until': pcm_projected,
                    'timesheets_approved_until': ts_until,
                    'notes': notes,
                }
            )

            if not created:
                # Update all fields so re-running the import picks up changes
                obj.name = name or obj.name
                obj.employee_number = employee_num or obj.employee_number
                obj.staff_type = staff_type or obj.staff_type
                obj.department = department or obj.department
                obj.pcm_until = pcm_until if pcm_until else obj.pcm_until
                obj.pcm_projected_until = pcm_projected if pcm_projected else obj.pcm_projected_until
                obj.timesheets_approved_until = ts_until if ts_until else obj.timesheets_approved_until
                obj.notes = notes or obj.notes
                obj.save()

            count += 1

        return count

    def _import_costs(self, wb):
        """Import salary/cost data from the 'costs' sheet. Unchanged from v1."""
        if 'costs' not in wb.sheetnames:
            return 0

        ws = wb['costs']
        count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            name     = safe_str(row[0])
            initials = safe_str(row[1])
            valid_from = safe_date(row[2])
            valid_until = safe_date(row[3])
            monthly = safe_float(row[4])

            if not initials or monthly is None:
                continue

            # Find the staff member by initials; create a minimal record if missing
            try:
                staff = StaffMember.objects.get(initials=initials)
            except StaffMember.DoesNotExist:
                staff, _ = StaffMember.objects.get_or_create(
                    initials=initials,
                    defaults={'name': name}
                )

            StaffCost.objects.get_or_create(
                staff_member=staff,
                valid_from=valid_from or datetime.date(2022, 1, 1),
                defaults={
                    'valid_until': valid_until or datetime.date(2030, 12, 31),
                    'monthly_cost': monthly,
                    'confidence': safe_str(row[5]),
                    'notes': safe_str(row[6]),
                }
            )
            count += 1

        return count

    def _import_projects(self, wb):
        """
        Import projects from the 'projects' sheet.

        The column positions for dates and metadata are the same as v1.
        The staff columns in the middle have changed (AF and PS added,
        HI1 and SRSE1 removed) but we don't read those individual columns —
        we read allocations from the per-person sheets instead.
        """
        if 'projects' not in wb.sheetnames:
            return 0

        ws = wb['projects']
        count = 0

        def to_bool(val):
            if val == 'Yes':
                return True
            if val == 'No':
                return False
            return None

        # Data starts at row 4 (rows 1-3 are headers/labels)
        for row in ws.iter_rows(min_row=4, values_only=True):
            name = safe_str(row[0])
            if not name:
                continue

            rcode    = safe_str(row[1])
            is_ts    = to_bool(safe_str(row[2]))
            is_audit = to_bool(safe_str(row[3]))
            pm_orig  = safe_float(row[4])
            pm_add   = safe_float(row[5])
            pm_used  = safe_float(row[6])

            # Date columns — positions are the same as v1 regardless of staff column changes
            # because the date columns sit after all the staff columns at the far right.
            # With the new header the dates are at indices 39 (Start) and 40 (End).
            # We use -6, -5 etc. (from the end) to be robust to column count changes.
            start_date    = safe_date(row[-6])   # 'Start' column
            end_date      = safe_date(row[-5])   # 'End' column
            work_type     = safe_str(row[-3])    # 'type of work'
            possible_staff = safe_str(row[-2])   # 'possible staff'
            notes         = safe_str(row[-1])    # 'Notes'

            today = datetime.date.today()
            is_active = (
                True if (start_date and end_date and start_date <= today <= end_date)
                else None
            )

            obj, created = Project.objects.get_or_create(
                name=name,
                defaults={
                    'rcode': rcode,
                    'is_ts_required': is_ts,
                    'is_audited': is_audit,
                    'is_active': is_active,
                    'pm_original': pm_orig,
                    'pm_additional': pm_add,
                    'pm_used': pm_used,
                    'start_date': start_date,
                    'end_date': end_date,
                    'work_type': work_type,
                    'possible_staff': possible_staff,
                    'notes': notes,
                }
            )

            if not created:
                # Update fields that may have changed
                if rcode:
                    obj.rcode = rcode
                if start_date:
                    obj.start_date = start_date
                if end_date:
                    obj.end_date = end_date
                if pm_orig is not None:
                    obj.pm_original = pm_orig
                obj.notes = notes or obj.notes
                obj.save()

            count += 1

        return count

    def _import_project_budgets(self, wb):
        """Import project budget figures. Unchanged from v1."""
        if 'project_costs' not in wb.sheetnames:
            return 0

        ws = wb['project_costs']

        for row in ws.iter_rows(min_row=4, values_only=True):
            name = safe_str(row[0])
            if not name:
                continue

            try:
                project = Project.objects.get(name=name)
            except Project.DoesNotExist:
                continue

            di         = safe_float(row[4])
            da         = safe_float(row[5])
            additional = safe_float(row[6])

            ProjectBudget.objects.update_or_create(
                project=project,
                defaults={
                    'budget_di_original': di,
                    'budget_da_original': da,
                    'budget_additional': additional,
                }
            )

    def _import_work_packages(self, wb):
        """Import work packages. Unchanged from v1."""
        if 'project_WPs' not in wb.sheetnames:
            return 0

        ws = wb['project_WPs']
        count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            project_name = safe_str(row[0])
            wp_num       = safe_str(row[2])
            if not project_name or not wp_num:
                continue

            try:
                project = Project.objects.get(name=project_name)
            except Project.DoesNotExist:
                continue

            WorkPackage.objects.get_or_create(
                project=project,
                wp_number=wp_num,
                defaults={
                    'person_months': safe_float(row[3]),
                    'start_date': safe_date(row[6]),
                    'end_date': safe_date(row[7]),
                    'title': safe_str(row[8]),
                }
            )
            count += 1

        return count

    def _import_allocations(self, wb):
        """
        Import staff allocations from each person's individual sheet.

        The per-person sheet structure is the same as v1:
          Col B (index 1): From date
          Col C (index 2): Until date
          Col D (index 3): R code
          Col E (index 4): Project name
          Col F (index 5): Person-months
          Col G (index 6): FTE
          Col H (index 7): Hours/month
          Col I (index 8): Number of months
          Col J (index 9): Cost (£)

        The only change is PERSON_SHEETS at the top of this file, which now
        includes AF and PS and excludes HI1 and SRSE1.
        """
        count = 0

        for sheet_name in PERSON_SHEETS:
            if sheet_name not in wb.sheetnames:
                # Log a warning if a sheet we expect is missing — useful for debugging
                self.stdout.write(
                    self.style.WARNING(f"  Sheet '{sheet_name}' not found — skipping")
                )
                continue

            ws = wb[sheet_name]

            try:
                staff = StaffMember.objects.get(initials=sheet_name)
            except StaffMember.DoesNotExist:
                self.stdout.write(
                    self.style.WARNING(
                        f"  No StaffMember with initials '{sheet_name}' — "
                        f"run import again after staff sheet has been imported"
                    )
                )
                continue

            for row in ws.iter_rows(min_row=3, values_only=True):
                project_name = safe_str(row[4])
                if not project_name:
                    continue

                # Skip rows that are notes/headers rather than real allocation data
                skip_prefixes = ('Project', 'PCM line', 'Person', 'WP')
                if any(project_name.startswith(p) for p in skip_prefixes):
                    continue
                if project_name.startswith('('):
                    continue

                start_date = safe_date(row[1])
                end_date   = safe_date(row[2])
                if not start_date or not end_date:
                    continue

                fte = safe_float(row[6])
                if fte is None:
                    continue

                person_months   = safe_float(row[5])
                hours_per_month = safe_float(row[7])
                cost            = safe_float(row[9])
                notes           = safe_str(row[11]) if len(row) > 11 else ''

                # Find or create the project
                project, _ = Project.objects.get_or_create(
                    name=project_name,
                    defaults={'notes': 'Auto-created from allocation import'}
                )

                StaffAllocation.objects.get_or_create(
                    staff_member=staff,
                    project=project,
                    start_date=start_date,
                    end_date=end_date,
                    defaults={
                        'fte': fte,
                        'person_months': person_months,
                        'hours_per_month': hours_per_month,
                        'cost': cost,
                        'notes': notes,
                    }
                )
                count += 1

        return count

    def _import_swaps(self, wb):
        """Import budget swaps. Unchanged from v1."""
        if 'swapsies' not in wb.sheetnames:
            return 0

        ws = wb['swapsies']
        count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            from_name = safe_str(row[0])
            to_name   = safe_str(row[1])
            if not from_name or not to_name:
                continue

            swap_date = safe_date(row[2])
            amount    = safe_float(row[3])
            notes     = safe_str(row[4])
            action    = safe_str(row[5])

            from_project = None
            to_project   = None
            try:
                from_project = Project.objects.get(name__icontains=from_name)
            except (Project.DoesNotExist, Project.MultipleObjectsReturned):
                pass
            try:
                to_project = Project.objects.get(name__icontains=to_name)
            except (Project.DoesNotExist, Project.MultipleObjectsReturned):
                pass

            BudgetSwap.objects.get_or_create(
                from_project_name=from_name,
                to_project_name=to_name,
                amount_gbp=amount or 0,
                defaults={
                    'from_project': from_project,
                    'to_project': to_project,
                    'date': swap_date,
                    'notes': notes,
                    'action_to_balance': action,
                }
            )
            count += 1

        return count
